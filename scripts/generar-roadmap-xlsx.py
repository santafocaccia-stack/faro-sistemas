# -*- coding: utf-8 -*-
"""
Genera/regenera el roadmap de Gesto en Excel (roadmap-gesto.xlsx).
Fuente de verdad de tareas: lista TAREAS de abajo. Para actualizar el estado de
una tarea, editar acá y volver a correr, o editar el .xlsx directamente.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

OUT = r"C:\Users\Tomi\Documents\Software\faro-sistemas\roadmap-gesto.xlsx"
HOY = "2026-06-25"

# ---- Paleta "Brasas" ----
NARANJA   = "C2410C"  # acento
NARANJA_2 = "EA580C"
OSCURO    = "1C1917"
HEADER_BG = "1C1917"
HEADER_FG = "FFFFFF"
VERDE     = "DCFCE7"  # hecho
AMARILLO  = "FEF9C3"  # en progreso
GRIS      = "F1F5F9"  # pendiente
ZEBRA     = "FAFAF9"

# (categoria, tarea, prioridad, estado, notas)
# estado: "Hecho" | "En progreso" | "Pendiente"
TAREAS = [
    # Infraestructura
    ("Infraestructura", "Stack Next.js 16 App Router + TypeScript strict + Turbopack", "—", "Hecho", ""),
    ("Infraestructura", "Tailwind v4 + shadcn/ui + Design System Brasas", "—", "Hecho", ""),
    ("Infraestructura", "Drizzle ORM + Supabase Postgres", "—", "Hecho", ""),
    ("Infraestructura", "Supabase Auth SSR + middleware de sesión", "—", "Hecho", ""),
    ("Infraestructura", "Framer Motion + command palette (cmdk)", "—", "Hecho", ""),

    # Base de datos
    ("Base de datos", "Schema multi-tenant (tenants, users, productos, clientes, ventas, pagos, CC)", "—", "Hecho", "10 tablas (+ presupuestos)"),
    ("Base de datos", "Migraciones aplicadas en Supabase", "—", "Hecho", ""),
    ("Base de datos", "RLS multi-tenant (16 políticas, script defensivo)", "—", "Hecho", "byTenant() + RLS como red de seguridad"),

    # Auth
    ("Auth", "Login / Signup / Onboarding (crear tenant)", "—", "Hecho", ""),
    ("Auth", "Recuperar contraseña (forgot + reset)", "—", "Hecho", ""),
    ("Auth", "Auth callback PKCE + token_hash", "—", "Hecho", ""),

    # Productos
    ("Productos", "CRUD productos (stock kg/unidad, precios may/min)", "—", "Hecho", "Title Case, inputs vacíos no 0"),
    ("Productos", "Carga rápida (scan → nombre → precio → siguiente)", "—", "Hecho", "/productos/carga-rapida"),
    ("Productos", "Importación CSV con preview y validación", "—", "Hecho", "lotes de 100, crea categorías"),
    ("Productos", "Categorías (CRUD + filtro en lista y POS)", "—", "Hecho", ""),
    ("Productos", "Grupos de variantes + sugerencia asistida", "—", "Hecho", ""),
    ("Productos", "Código PLU (plan balanza)", "—", "Hecho", ""),
    ("Productos", "Precios vivos (anti-inflación: margen + actualización masiva)", "Alta", "En progreso", "Falta QA en vivo los 3 planes; ver docs/context/precios.md"),
    ("Productos", "Alertas de stock bajo (banner/email cuando stock ≤ mínimo)", "Media", "Pendiente", "El campo existe, no dispara nada"),
    ("Productos", "Precios especiales por cliente", "Media", "Pendiente", "Mayoristas con acuerdos puntuales"),

    # Clientes
    ("Clientes", "CRUD clientes + ficha con saldo CC", "—", "Hecho", "tipo, condición IVA, descuento"),

    # Ventas / POS
    ("Ventas / POS", "POS unificado mayorista + minorista (carrito Zustand)", "—", "Hecho", "split-screen"),
    ("Ventas / POS", "Scanner de código de barras (ZXing, cámara real)", "—", "Hecho", ""),
    ("Ventas / POS", "Métodos de pago (efectivo, transf, tarjeta, MP, cuenta corriente)", "—", "Hecho", ""),
    ("Ventas / POS", "Ticket 80mm + Remito A4 PDF (compartir / descargar)", "—", "Hecho", "@react-pdf/renderer"),
    ("Ventas / POS", "Atajos de teclado (F2 cobrar, Esc, Supr)", "—", "Hecho", ""),
    ("Ventas / POS", "Historial + detalle de venta (filtro por período)", "—", "Hecho", ""),
    ("Ventas / POS", "Descuentos en POS (% o $ por total)", "Media", "Pendiente", "Muy pedido en kioscos/carnicerías"),
    ("Ventas / POS", "Devoluciones / notas de crédito (reversar venta + ajuste stock)", "Media", "Pendiente", "Reversión de CC ya está hecha"),
    ("Ventas / POS", "Lector de códigos USB en POS desktop (Enter agrega el producto escaneado)", "—", "Hecho", "Minorista + mayorista; el lector tipo supermercado escribe el código y Enter lo suma al carrito"),

    # Cuenta Corriente
    ("Cuenta Corriente", "Movimientos + registrar pago + saldo en tiempo real", "—", "Hecho", "registrarPago atómico FOR UPDATE"),
    ("Cuenta Corriente", "Reversión de movimientos (transacción atómica)", "—", "Hecho", "motivo obligatorio"),

    # Reportes
    ("Reportes", "KPIs por período + ranking top productos + métodos de pago", "—", "Hecho", "ventana de tiempo unificada"),
    ("Reportes", "Exportar a CSV", "—", "Hecho", ""),

    # Presupuestos
    ("Presupuestos", "CRUD + PDF + estados + numeración atómica", "—", "Hecho", "db.transaction"),
    ("Presupuestos", "Plantillas de servicios (guardar/listar/usar)", "—", "Hecho", ""),

    # Proveedores / Pedidos
    ("Proveedores / Pedidos", "CRUD proveedores (markup, selector de país)", "—", "Hecho", ""),
    ("Proveedores / Pedidos", "Pedidos a proveedor (WhatsApp + recepción con ajuste de stock)", "—", "Hecho", ""),

    # Configuración / Equipo
    ("Configuración / Equipo", "Datos fiscales + canales + conectar/desconectar Mercado Pago", "—", "Hecho", ""),
    ("Configuración / Equipo", "Equipo: invitar por email, editar rol, eliminar", "—", "Hecho", ""),
    ("Configuración / Equipo", "Control de acceso por rol (requireAdmin, route group)", "—", "Hecho", "empleado solo POS+Historial+Productos"),

    # UX / Diseño
    ("UX / Diseño", "Design System Brasas v2 (paleta cálida + grain + gradientes)", "—", "Hecho", ""),
    ("UX / Diseño", "Sidebar + mobile bottom nav con FAB Vender", "—", "Hecho", ""),
    ("UX / Diseño", "Skeleton loaders (7 rutas)", "—", "Hecho", ""),
    ("UX / Diseño", "Command palette ⌘K", "—", "Hecho", ""),
    ("UX / Diseño", "Nomenclatura: los modos se presentan como 'versión' al usuario (Market, Servicios, etc.)", "—", "Hecho", "Landing (switcher, precios, cierre) + /planes + confirmación de transferencia; se usan los nombres reales de los planes (Market, no inventados como 'Kiosco')"),
    ("UX / Diseño", "Dark / light toggle", "Baja", "Pendiente", ""),

    # Negocio SaaS / Suscripción
    ("Negocio SaaS", "4 planes + Mercado Pago preapproval + webhook activación/suspensión", "—", "Hecho", "src/lib/planes.ts"),
    ("Negocio SaaS", "Cobro por transferencia (panel admin + aviso cliente + comprobante por email)", "—", "Hecho", "Activación manual 1-click; /admin/suscripciones; emails salen al cargar RESEND_API_KEY"),
    ("Negocio SaaS", "Cobro automático MP/Mobbex (CUIT Monotributo + cuenta)", "Alta", "En progreso", "Bloqueado: falta CUIT + cuenta. Mientras tanto se cobra por transferencia"),
    ("Negocio SaaS", "Panel super-admin (gestionar tenants, métricas cross-tenant)", "Media", "En progreso", "Panel /admin/suscripciones operativo; faltan métricas cross-tenant"),
    ("Negocio SaaS", "Catálogo global cross-tenant", "Baja", "Pendiente", "Necesita OK explícito antes de implementar"),

    # Lanzamiento (bloqueantes)
    ("Lanzamiento (bloqueante)", "Dominio propio (ej: usegesto.app)", "Alta", "Pendiente", "URL actual no es presentable"),
    ("Lanzamiento (bloqueante)", "SMTP con Resend (invitación, bienvenida, reset con branding)", "Alta", "Pendiente", "Cron listo; falta RESEND_API_KEY en Vercel (rotar la vieja)"),
    ("Lanzamiento (bloqueante)", "Landing page pública (adaptativa por versión + demo en vivo)", "Alta", "Hecho", "v3 'de oficio' LIVE: switcher de versiones auto-rotativo, demo POS con ticket de papel, foto real + captura de Gesto en celular, antes/después, FAQ, pricing al MEP, paleta terracota (sin neón), responsive mobile. Login rediseñado (split de marca)"),
    ("Lanzamiento (bloqueante)", "PWA manifest + iconos (192/512, apple-touch-icon)", "Alta", "Pendiente", "POS se usa en tablet/celular"),
    ("Lanzamiento (bloqueante)", "Términos de servicio + Política de privacidad", "Alta", "Hecho", "Páginas /legal/{terminos,privacidad,cookies,baja-de-datos} + links en footer + checkbox en signup. Prestador: Faro Sistemas (SIN datos fiscales por etapa pre-formalización; único contacto = email en constante CONTACTO_EMAIL, hoy el gmail). AL FORMALIZAR: cargar CUIT + domicilio + jurisdicción y que un abogado revise"),

    # Calidad / Seguridad
    ("Calidad / Seguridad", "Bugs críticos corregidos (stock al vender, numeración, loop /planes)", "—", "Hecho", ""),
    ("Calidad / Seguridad", "ConfirmDialog reutilizable (reemplaza window.confirm en 9 comp.)", "—", "Hecho", ""),
    ("Calidad / Seguridad", "Verificar firma de webhook MP", "—", "Hecho", "Firma HMAC timing-safe + fail-closed en prod; cargar MP_WEBHOOK_SECRET en Vercel"),
    ("Calidad / Seguridad", "Auditoría de seguridad E2E + fixes (IDOR token MP, recibirPedido, escalación de rol, OAuth state, crons fail-closed, escape de emails)", "—", "Hecho", "Guard test anti-omisión de byTenant; ver project_gesto_seguridad. RLS no protege capa app (byTenant única defensa)"),
    ("Calidad / Seguridad", "Rate limiting en auth (login/signup)", "Media", "Pendiente", "ej: @upstash/ratelimit"),
    ("Calidad / Seguridad", "Tests de integración (flows de venta y producto)", "Media", "Pendiente", ""),
    ("Calidad / Seguridad", "Export de datos (CSV ventas / inventario)", "Media", "Pendiente", "Que el cliente no sienta lock-in"),
    ("Calidad / Seguridad", "Facturación AFIP (Tusfacturas.app, A/B)", "Media", "Pendiente", "Bloqueante segmento servicios/mayorista"),
    ("Calidad / Seguridad", "Historial de auditoría básico (quién creó/modificó)", "Baja", "Pendiente", "Mínimo en productos y ventas"),

    # Legal / Fiscal (Argentina) — para cobrar en regla y escalar
    ("Legal / Fiscal", "CUIT + Monotributo (o Responsable Inscripto)", "Alta", "Pendiente", "Bloquea cobrar en regla; ojo el tope anual de Monotributo con ingresos SaaS recurrentes"),
    ("Legal / Fiscal", "Registro de bases de datos ante la AAIP (Ley 25.326)", "Media", "Pendiente", "Sos responsable de la PII de los clientes y de SUS clientes"),
    ("Legal / Fiscal", "Botón de baja + arrepentimiento de suscripción (Res. 424/2020)", "Media", "En progreso", "Ya existe cancelarSuscripcion + página /legal/baja-de-datos que explica el proceso; falta el botón de arrepentimiento bien visible"),
    ("Legal / Fiscal", "Cláusulas de transferencia internacional de datos", "Baja", "Pendiente", "Datos en AWS us-east-1; Ley 25.326 exige resguardos (EE.UU. no es país adecuado por defecto)"),
    ("Legal / Fiscal", "Estructura societaria (SAS) + acuerdo escrito con socio", "Media", "Pendiente", "Separa patrimonio del riesgo; formaliza la participación de Facu"),
    ("Legal / Fiscal", "Contrato de servicio / SLA con clientes", "Media", "Pendiente", "Dueño de los datos (el cliente), uptime, soporte, límite de responsabilidad"),

    # Infra / Escalado
    ("Infra / Escalado", "Vercel Pro + Supabase Pro", "Alta", "Pendiente", "Vercel Hobby es no comercial; Supabase Free se pausa a los 7 días. Bloquea producción real"),
    ("Infra / Escalado", "Backups diarios + PITR + probar un restore", "Media", "Pendiente", "Un backup sin restore probado no es un backup"),
    ("Infra / Escalado", "Monitoreo: activar Sentry + uptime check", "Media", "Pendiente", "SENTRY_DSN ya está en el build, falta activarlo + un cron canario"),
    ("Infra / Escalado", "Email deliverability: dominio + SPF/DKIM/DMARC (Resend)", "Media", "Pendiente", "Si no, comprobantes y facturas caen en spam"),
    ("Infra / Escalado", "Encriptar tokens MP del negocio at-rest", "Media", "Pendiente", "Hoy en texto plano en tenants (hallazgo de seguridad #9)"),

    # Backlog destacado
    ("Backlog / Ideas", "Consultor de precios para cliente final (pantalla pública con QR)", "Media", "Pendiente", "Diferenciador único; ruta /p/[tenantSlug]"),

    # Versiones / Verticales (mejoras por versión — feedback del dueño 2026-06-21)
    ("Versiones / Verticales", "Mejorar login/registro: sumarle contenido útil (hoy se siente vacío)", "Media", "Hecho", "Split de marca + recuperación de contraseña ya estaban. Sumado: ver/ocultar contraseña (PasswordInput) en login/signup/reset, errores específicos (email sin confirmar, rate-limit, email ya registrado), pie con Términos/Privacidad/ayuda por mail, indicador de fuerza de contraseña"),
    ("Versiones / Verticales", "Gesto Balanza funcional: integración con balanza digital + venta por peso", "Alta", "Hecho", "Plan habilitado + nav + página guía /dashboard/balanza. Lectura de etiquetas de balanza con peso embebido (EAN-13 prefijo 2 -> PLU + peso) PORTADA al POS real (pos-container; el primer intento quedó en un componente muerto ya borrado). Peso a mano: modal de peso al tocar/escanear un producto por_kg (gramos + atajos 250/500/1000 g + subtotal en vivo). Pendiente futuro: configurar el formato exacto de etiqueta por marca de balanza"),
    ("POS / Hardware", "Impresión de tickets: térmica directa por Bluetooth y USB (ESC/POS)", "Alta", "Hecho", "Motor ESC/POS (CP850), transporte Web Bluetooth + WebUSB, botón 'Térmica' en el modal post-venta que conecta la ticketera e imprime de un toque. window.print() (sistema/AirPrint) y PDF ya existían. iOS solo por window.print(). Pendiente futuro: impresoras WiFi por IP (puerto 9100) requieren un puente/print-server o ePOS HTTP; app nativa abriría todo"),
    ("Versiones / Verticales", "Gesto Food: POS propio (comandas, mesas, cocina/KDS) — muy distinto a Market", "Alta", "Pendiente", "Hoy figura 'próximamente'"),
    ("Versiones / Verticales", "Gesto Servicios: emitir boletas además de presupuestos + simplificar botones/UI", "Alta", "Hecho", "Form de presupuesto simplificado + BOLETAS de punta a punta: boleta directa (form modo boleta con método de pago), numeración propia, lista /dashboard/boletas + nav, detalle adaptado y PDF como RECIBO. Migración 0009 aplicada en dev+prod. Comprobante NO fiscal. Botón 'Emitir recibo' desde un presupuesto cobrado HECHO (PDF como recibo, sin duplicar ingreso). Pendiente menor: seguir puliendo UI"),
    ("Versiones / Verticales", "Gesto Atmosféricos: historial, agenda y pedidos para días futuros + arreglar ruta en Maps", "Alta", "Hecho", "Historial ya existía; AGENDA (navegación entre días) + pedidos para días futuros HECHO. Maps ARREGLADO: campo 'ubicación exacta' (link de Maps) por pedido; geo.ts extrae coordenadas del link y la ruta usa el punto exacto en vez del texto. Migración 0010 en dev+prod"),

    # Roadmap v2
    ("Roadmap v2", "Modificadores por ítem en POS (extras, variantes)", "Baja", "Pendiente", ""),
    ("Roadmap v2", "App mobile nativa con Capacitor (offline first)", "Baja", "Pendiente", ""),
    ("Roadmap v2", "Notificaciones push (stock, pedidos)", "Baja", "Pendiente", ""),
    ("Roadmap v2", "Dashboard financiero avanzado (márgenes, rentabilidad)", "Baja", "Pendiente", ""),
    ("Roadmap v2", "Multi-sucursal (varios locales bajo una cuenta madre)", "Baja", "Pendiente", ""),

    # ── Marketing / Conversión (análisis de IAs externas) ──
    ("Marketing / Conversión", "Registro en mini-wizard modal (versión -> cuenta), sin mandar al login frío", "Alta", "Hecho", "Modal in-page (no popup) que abren los CTA 'Empezá gratis'. Paso 1 elegís versión, paso 2 creás cuenta; la versión se hereda al onboarding"),
    ("Marketing / Conversión", "Login con Google (OAuth)", "Alta", "En progreso", "Código cableado en el wizard y disponible. FALTA: habilitar el provider Google en Supabase (Auth -> Providers) con credenciales de Google Cloud Console"),
    ("Marketing / Conversión", "Botón de WhatsApp flotante en la landing", "Alta", "Hecho", "Sugerido por Deepseek y Gemini. Número real cargado (wa.me/5491166644837), mensaje pre-cargado"),
    ("Marketing / Conversión", "Landing pages por vertical (/market, /atmosfericos, /balanza) para pauta en Meta Ads", "Alta", "Pendiente", "Gemini: la página de destino tiene que hablarle solo a ese rubro para no diluir la conversión"),
    ("Marketing / Conversión", "Dominio propio (ej. gesto.com.ar) apuntado en Vercel", "Alta", "Pendiente", "Gemini: faro-sistemas.vercel.app no transmite 'Gesto' y resta confianza"),
    ("Marketing / Conversión", "Nav de la landing con Versiones y Ayuda (hoy solo Precios)", "Media", "Hecho", "Nav: 'Versiones' (antes 'Cómo funciona') + 'Ayuda' al WhatsApp"),
    ("Marketing / Conversión", "CTA por versión en precios (abre el wizard con esa versión)", "Media", "Hecho", "Cada tarjeta: 'Probar [versión] gratis' -> abre el wizard con esa versión preelegida"),
    ("Marketing / Conversión", "Precio en ARS más visible (manteniendo USD como principal)", "Media", "Hecho", "El ARS se agrandó (16px, seminegrita); el USD sigue siendo el número grande. Se mantiene el cobro al MEP en USD"),
    ("Marketing / Conversión", "Sello 'datos encriptados' + destacar '14 días sin tarjeta' como badge", "Media", "Hecho", "Sellos bajo los planes: datos encriptados (SSL), 14 días sin tarjeta, cancelás cuando quieras"),
    ("Marketing / Conversión", "Testimonios / casos de éxito", "Media", "Pendiente", "Dejar el espacio diseñado y llenarlo con los primeros clientes reales (no inventar)"),

    # ── QA por planes (jun-2026) — feedback del dueño probando cada versión ──
    ("QA por planes (jun-2026)", "MP: conectar cuenta del negocio (OAuth) — arreglado, conecta en prod", "Alta", "Hecho", "Faltaban MP_APP_ID/MP_CLIENT_SECRET en Vercel + redirect_uri canónico. Se fijó NEXT_PUBLIC_APP_URL=faro-sistemas (eliminado el fallback -gold) y mp-url usa getAppUrl(). Callback con diagnóstico ?reason=. Cuenta de prueba conectó OK (ID 642364293)"),
    ("QA por planes (jun-2026)", "Cobro REAL por Mercado Pago (QR dinámico / link / Point) con la cuenta conectada", "Alta", "Pendiente", "HOY 'Mercado Pago' en el POS solo MARCA la venta como pagada (registro manual). getMPNegocioToken existe pero NO se usa. Falta: crear QR/orden por API de MP + webhook de confirmación que marque la venta cobrada"),
    ("QA por planes (jun-2026)", "Nav: 'Equipo' ya no marca también 'Configuración'", "Media", "Hecho", "match por href más específico. En staging, falta merge a prod"),
    ("QA por planes (jun-2026)", "Nav PC: sacar el desplegable 'Más', listar todos los items directos", "Media", "Hecho", "staging"),
    ("QA por planes (jun-2026)", "POS: el hint 'escribí para buscar o elegí categoría' desaparece al enfocar el buscador", "Baja", "Hecho", "staging"),
    ("QA por planes (jun-2026)", "POS: cambiar cantidad en el carrito suena igual que agregar producto", "Baja", "Hecho", "staging"),
    ("QA por planes (jun-2026)", "POS: sonido de agregar más contundente (~1.4x, beepAgregar)", "Baja", "Hecho", "staging"),
    ("QA por planes (jun-2026)", "POS: código de barras del buscador más grande/legible", "Baja", "Hecho", "staging"),
    ("QA por planes (jun-2026)", "POS: fix ticket en blanco al imprimir (Market)", "Alta", "Hecho", "El ticket print-only vivía dentro de un contenedor .screen-only (display:none al imprimir); ahora va por portal a document.body. PDF ya andaba. staging"),
    ("QA por planes (jun-2026)", "Productos: stock muestra '6' no '6,000' (limpia ceros de relleno de Postgres)", "Media", "Hecho", "staging"),
    ("QA por planes (jun-2026)", "Carga rápida: agregar campo de stock inicial (con navegación por Enter)", "Media", "Hecho", "staging"),
    ("QA por planes (jun-2026)", "Ticket de cuenta corriente con resumen del saldo del cliente", "Media", "Pendiente", ""),
    ("QA por planes (jun-2026)", "Selector de tamaño de fuente global (accesibilidad, preview en vivo)", "Media", "Pendiente", "Para gente que ve poco. Seleccionable al inicio/onboarding + en config"),
    ("QA por planes (jun-2026)", "Eliminar grupos de variantes — dejar solo categorías (el buscador ya cubre)", "Media", "Pendiente", "DESTRUCTIVA: migración de DB. Reemplaza la entrada 'Grupos de variantes'. Revisar impacto en plan balanza"),
    ("QA por planes (jun-2026)", "Probar importación CSV (generar un CSV de ejemplo y validar)", "Media", "Pendiente", ""),
    ("QA por planes (jun-2026)", "Filtros en el historial (por fecha, cliente, etc.)", "Media", "Pendiente", "Hoy solo filtra por período"),
    ("QA por planes (jun-2026)", "Cliente: al seleccionarlo ver su ficha (datos/historial/saldo), no el form de edición; botón 'Editar' aparte", "Media", "Pendiente", ""),
    ("QA por planes (jun-2026)", "Reportes: al cambiar el período recalcular TODOS los datos, no solo el gráfico", "Media", "Pendiente", "Verificar; con pocas ventas en demo puede ocultarse"),
    ("QA por planes (jun-2026)", "CC: aclarar/mejorar UX de 'no permitir saldo negativo'", "Baja", "Pendiente", "El usuario no entiende cómo se maneja hoy"),
    ("QA por planes (jun-2026)", "Proveedor: revisar selector de bandera de país (no se ve al crear)", "Baja", "Pendiente", "Decidir si va y por qué"),
    ("QA por planes (jun-2026)", "Botón 'generar pedido' desde Proveedor y desde Pedidos", "Media", "Pendiente", "Hoy no hay forma de armar el pedido a proveedor"),
    ("QA por planes (jun-2026)", "Categorías de gastos faltan en plan Market (revisar todos los planes)", "Media", "Pendiente", ""),
    ("QA por planes (jun-2026)", "Balance mensual IA automático (cron fin de mes / fecha configurable), sin botón manual", "Media", "Pendiente", "Opción: último día hábil del mes o fecha elegida"),
    ("QA por planes (jun-2026)", "Permisos granulares por usuario (capacidades sobre el rol empleado)", "Alta", "Pendiente", "users_tenants.permisos ya existe pero sin uso. Modelo elegido: toggles por capacidad (POS, Productos, Clientes, Reportes, Gastos, etc.)"),
]

# ================= Construcción del libro =================
wb = Workbook()
ws = wb.active
ws.title = "Roadmap"

thin = Side(style="thin", color="E2E0DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Título
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "Roadmap — Gesto (Faro Sistemas)"
c.font = Font(name="Calibri", size=18, bold=True, color=NARANJA)
c.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = f"SaaS de gestión comercial multi-tenant · Actualizado {HOY} · Editá la columna Estado y la fecha al cerrar cada tarea"
c.font = Font(size=10, italic=True, color="78716C")

headers = ["ID", "Categoría", "Tarea", "Prioridad", "Estado", "Notas", "Actualizado"]
HEADER_ROW = 4
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=HEADER_ROW, column=col, value=h)
    cell.font = Font(bold=True, color=HEADER_FG, size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
ws.row_dimensions[HEADER_ROW].height = 22

estado_fill = {
    "Hecho": PatternFill("solid", fgColor=VERDE),
    "En progreso": PatternFill("solid", fgColor=AMARILLO),
    "Pendiente": PatternFill("solid", fgColor=GRIS),
}

r = HEADER_ROW + 1
for i, (cat, tarea, prio, estado, notas) in enumerate(TAREAS, start=1):
    fila = [i, cat, tarea, prio, estado, notas, HOY]
    for col, val in enumerate(fila, start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(col in (3, 6)))
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if col == 4:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if col == 5:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
    r += 1

LAST = r - 1

# Anchos
widths = {"A": 5, "B": 22, "C": 56, "D": 11, "E": 13, "F": 46, "G": 13}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Congelar y filtro
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{HEADER_ROW}:G{LAST}"

# Validación de datos (dropdowns) en Estado y Prioridad
dv_estado = DataValidation(type="list", formula1='"Hecho,En progreso,Pendiente"', allow_blank=False)
dv_prio = DataValidation(type="list", formula1='"Alta,Media,Baja,—"', allow_blank=True)
ws.add_data_validation(dv_estado)
ws.add_data_validation(dv_prio)
dv_estado.add(f"E5:E{LAST}")
dv_prio.add(f"D5:D{LAST}")

# Formato condicional por Estado (colorea toda la fila A:G)
rng = f"A5:G{LAST}"
ws.conditional_formatting.add(rng, FormulaRule(formula=['$E5="Hecho"'], fill=estado_fill["Hecho"]))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$E5="En progreso"'], fill=estado_fill["En progreso"]))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$E5="Pendiente"'], fill=estado_fill["Pendiente"]))

# ================= Hoja Resumen =================
ws2 = wb.create_sheet("Resumen")
ws2.merge_cells("A1:C1")
c = ws2["A1"]
c.value = "Resumen del Roadmap"
c.font = Font(size=16, bold=True, color=NARANJA)
ws2.row_dimensions[1].height = 26

# Por estado (con fórmulas, se actualiza solo)
ws2["A3"] = "Estado"
ws2["B3"] = "Cantidad"
ws2["C3"] = "% del total"
for col in ("A3", "B3", "C3"):
    ws2[col].font = Font(bold=True, color=HEADER_FG)
    ws2[col].fill = PatternFill("solid", fgColor=HEADER_BG)
    ws2[col].alignment = Alignment(horizontal="center")

estados = ["Hecho", "En progreso", "Pendiente"]
total_ref = f"Roadmap!$E$5:$E${LAST}"
for idx, e in enumerate(estados):
    rr = 4 + idx
    ws2.cell(row=rr, column=1, value=e).fill = estado_fill[e]
    ws2.cell(row=rr, column=1).font = Font(bold=True)
    ws2.cell(row=rr, column=2, value=f'=COUNTIF({total_ref},A{rr})')
    ws2.cell(row=rr, column=3, value=f'=IFERROR(B{rr}/$B$7,0)').number_format = "0%"
ws2["A7"] = "TOTAL"
ws2["A7"].font = Font(bold=True)
ws2["B7"] = f'=COUNTA({total_ref})'
ws2["B7"].font = Font(bold=True)

# Por categoría
ws2["A10"] = "Categoría"
ws2["B10"] = "Hecho"
ws2["C10"] = "Pendiente/En curso"
for col in ("A10", "B10", "C10"):
    ws2[col].font = Font(bold=True, color=HEADER_FG)
    ws2[col].fill = PatternFill("solid", fgColor=HEADER_BG)
    ws2[col].alignment = Alignment(horizontal="center")

cats = []
for (cat, *_rest) in TAREAS:
    if cat not in cats:
        cats.append(cat)
cat_ref = f"Roadmap!$B$5:$B${LAST}"
for idx, cat in enumerate(cats):
    rr = 11 + idx
    ws2.cell(row=rr, column=1, value=cat)
    ws2.cell(row=rr, column=2,
             value=f'=COUNTIFS({cat_ref},A{rr},{total_ref},"Hecho")')
    ws2.cell(row=rr, column=3,
             value=f'=COUNTIFS({cat_ref},A{rr})-B{rr}')

ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 20

wb.save(OUT)
print("OK ->", OUT, "| filas:", len(TAREAS))
